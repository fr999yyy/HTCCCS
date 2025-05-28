from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User, auth
from django.contrib import messages
from .auth_utils import std_authenticate, std_login, std_logout, get_student
import zipfile, csv, shutil, os
import random
from django.core.files.storage import FileSystemStorage
from django.db import connection
from .models import Student, Section, Volunteer, Course, AdminSetting, Selection, SpecialCourse, SelectionResult, CustomUser
from django.contrib.contenttypes.models import ContentType
import pandas as pd
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.decorators import user_passes_test
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from django.db import transaction



# Create your views here.



def index(request):
    # user = request.user
    # if not user.is_anonymous:
    #     print(user)
    #     if user.has_perm('django_app.is_student'):
    #         return redirect('/select_form')
    #     elif user.has_perm('django_app.is_volunteer'):
    #         return redirect('/volunteer_dashboard')
    #     elif user.has_perm('django_app.is_cs'):
    #         return redirect('/dashboard')
    #     elif user.has_perm('django_app.is_admin'):
    #         return redirect('/dashboard')
    # else:
    #     print('anonymous')
    return render(request, 'index.html')    

def is_admin(user):
    return user.is_superuser

def stdLogin(request): # 學生登入
    if request.method == 'POST':
        if not request.POST['std_id'] or not request.POST['std_name']:
            messages.error(request, '請填寫所有欄位')
            return redirect('/stdLogin')
        std_id = request.POST['std_id']
        std_name = request.POST['std_name']
        # team = request.POST['team']
        # satb = request.POST['satb']
        student = std_authenticate(std_id=std_id, std_name=std_name)
        if student is not None:
            std_login(request, student)
            print('login success')
            return redirect ('std_index')
            pass
        else:
            messages.info(request, '無法登入，請檢查資料是否正確')
            print('login failed')
            return redirect('/stdLogin')
            pass
    else:
        return render(request, 'stdLogin.html', {'names': list(Student.objects.values_list('std_name', flat=True))})

def csLogin(request): # 選課組登入
    if request.method == 'POST':
        if not request.POST['password']:
            messages.error(request, '請輸入密碼')
            return redirect('/csLogin')
        password = request.POST['password']

        user = auth.authenticate(username='admin', password=password)
        if user is None:
            user = auth.authenticate(username='course_selection', password=password)

        if user is not None:
            auth.login(request, user)
            return redirect('/dashboard')
        else:
            messages.error(request, '密碼錯誤，請重新輸入')
            return redirect('/csLogin')
    else:
        return render(request, 'csLogin.html')

def Volunteer_Login(request): # 志工登入
    if request.method == 'POST':
        if not request.POST['password']:
            messages.error(request, '請輸入密碼')
            return redirect('/Volunteer_Login')
        password = request.POST['password']

        user = auth.authenticate(username='volunteer', password=password)

        if user is not None:
            auth.login(request, user)
            print('login success')
            return redirect('volunteer_dashboard')
        else:
            messages.error(request, '密碼錯誤，請重新輸入')
            print('pwd wrong')
            return redirect('/Volunteer_Login')
    else:
        print('login failed')
        return render(request, 'Volunteer_Login.html')

def volunteer_dashboard(request): # 志工主頁
    return render(request, 'volunteer_dashboard.html', {'sections': Section.objects.all()})

def dashboard(request): # 選課組主頁
    if request.user.is_authenticated:
        return render(request, 'dashboard.html')
    else:
        print('not authenticated')
        return redirect('/csLogin')

def courses_lookup(request): # 從資料庫抓課程回傳到前端
    section_id = request.GET.get('section_id') if request.GET.get('section_id') else 'all'
    if section_id == 'all':
        return JsonResponse({'courses': list(Course.objects.all().values('course_id', 'course_name')) + list(SpecialCourse.objects.all().values('course_id', 'course_name'))})
    courses_in_section = list(Course.objects.filter(section_id=section_id).values('course_id', 'course_name'))
    special_courses_in_section = list(SpecialCourse.objects.filter(section_id=section_id).values('course_id', 'course_name'))
    all_courses_in_section = courses_in_section + special_courses_in_section

    return JsonResponse({'courses': all_courses_in_section})

def selection_lookup(request): # 選課組後台-查詢選課結果
    if request.method == 'POST':
        input_course_id = request.POST['input_course_id']
        output_course_id = request.POST['output_course_id']

        students = SelectionResult.objects.filter(object_id=input_course_id).select_related('std')
        selections = Selection.objects.filter(course_id=output_course_id, std__in=[student.std_id for student in students]).order_by('priority')
        
        selection_details = []
        for selection in selections:
            student_name = Student.objects.get(std_id=selection.std_id).std_name
            selection_details.append({
                'priority': selection.priority,
                'student_name': student_name
            })

        return JsonResponse({'selections': selection_details})

    return render(request, 'selection_lookup.html', {'sections': Section.objects.all()})

def team_results_lookup(request): # 小組選課結果查詢
    if request.method == 'POST':
        j_or_h = request.POST.get('j_or_h')
        team = request.POST.get('team')
        section_id = request.POST.get('section_id')

        students = Student.objects.filter(j_or_h=j_or_h, team=team)

        results = []
        for student in students:
            result = SelectionResult.objects.filter(std=student, section_id=section_id).first()
            if result:
                results.append({
                    'std_name': student.std_name,
                    'course_name': result.course.course_name
                })

        return JsonResponse({'results': results})

    return render(request, 'volunteer_dashboard.html', {'sections': Section.objects.all()})

def course_results_lookup(request): # 課程選課結果查詢
    if request.method == 'POST':
        course_id = request.POST.get('course_id')
        # Get SelectionResult with the corresponding course_id
        results = []
        selection_results = SelectionResult.objects.filter(object_id=course_id)
        for result in selection_results:
            results.append({
                    'std_name': result.std.std_name,
                    'course_name': result.course.course_name,
                    'std_tag': result.std.std_tag
                })
            print(result.std.std_tag)

        return JsonResponse({'results': results})

    return render(request, 'volunteer_dashboard.html', {'sections': Section.objects.all(), 'courses': Course.objects.all()})

def upload_result_change(request): # 選課組後台-志願結果更改
    if request.method == 'POST' and request.FILES.get('upload_result_change'):
        result_change_file = request.FILES['upload_result_change']
        wb = openpyxl.load_workbook(result_change_file)
        sheet = wb.active

        changes = []
        errors = False

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Starting from row 2 (after header)
            # Skip empty rows
            if not any(row):
                continue
                
            # Check if row has enough columns
            if len(row) < 3:
                messages.error(request, f"Row {row_idx} doesn't have enough columns. Expected: section_id, student_name, course_name")
                errors = True
                continue
                
            section_id, student_name, course_name = row
            
            # Check for None values
            if section_id is None or student_name is None or course_name is None:
                messages.error(request, f"Row {row_idx} contains empty values: section_id={section_id}, student_name={student_name}, course_name={course_name}")
                errors = True
                continue
    
            try:
                student = Student.objects.get(std_name=student_name)
                section = Section.objects.get(section_id=section_id)
                selection_result = SelectionResult.objects.get(std=student, section=section)
    
                # Get the course_id from Course or SpecialCourse
                course = Course.objects.filter(course_name=course_name).first()
                if not course:
                    course = SpecialCourse.objects.filter(course_name=course_name).first()
    
                if course:
                    changes.append({
                        'selection_result': selection_result,
                        'course': course
                    })
                else:
                    messages.error(request, f"Course '{course_name}' not found for student '{student_name}' in section '{section_id}'.")
                    errors = True
    
            except Student.DoesNotExist:
                messages.error(request, f"Student '{student_name}' not found.")
                errors = True
            except Section.DoesNotExist:
                messages.error(request, f"Section '{section_id}' not found.")
                errors = True
            except SelectionResult.DoesNotExist:
                messages.error(request, f"SelectionResult not found for student '{student_name}' in section '{section_id}'.")
                errors = True

        if not errors and changes:
            # Apply changes
            with transaction.atomic():  # Use a transaction to ensure all changes are applied or none
                for change in changes:
                    selection_result = change['selection_result']
                    course = change['course']
                    
                    # Update content_type and object_id based on the course type
                    if isinstance(course, SpecialCourse):
                        content_type = ContentType.objects.get_for_model(SpecialCourse)
                    else:
                        content_type = ContentType.objects.get_for_model(Course)
                    
                    selection_result.content_type = content_type
                    selection_result.object_id = course.pk
                    selection_result.save()
                
                messages.success(request, f"Successfully updated {len(changes)} selection results.")
        elif not changes and not errors:
            messages.warning(request, "No changes were made. The file might be empty or contain no valid changes.")

        return redirect('updateData')
    
    return render(request, 'dashboard.html')


def result(request): # 選課組後台-志願結果頁面
    # Get selection stage range for section ID comparison
    selection_range = AdminSetting.objects.get(setting_name='J1stRange').configuration

    # Find all students who have submitted selections for the first form (sections 1-6)
    students_with_form1 = Student.objects.filter(
        std_id__in=Selection.objects.filter(
            section_id__lte=selection_range
        ).values_list('std_id', flat=True).distinct()
    )

    # Mark these students as having completed form1
    students_with_form1.update(form1_completed=True)

    # Find all students who have submitted selections for the second form (sections 7+)
    students_with_form2 = Student.objects.filter(
        std_id__in=Selection.objects.filter(
            section_id__gt=selection_range
        ).values_list('std_id', flat=True).distinct()
    )

    # Mark these students as having completed form2
    students_with_form2.update(form2_completed=True)
    
    # Get students who haven't completed form1
    incomplete_form1_students = Student.objects.filter(form1_completed=False)
    
    # Get students who haven't completed form2
    incomplete_form2_students = Student.objects.filter(form2_completed=False)
    
    context = {
        'incomplete_form1_students': incomplete_form1_students,
        'incomplete_form2_students': incomplete_form2_students,
    }
    return render(request, 'result.html', context)

def updateData(request):
    J1stRange = AdminSetting.objects.get(setting_name='J1stRange').configuration
    H1stRange = AdminSetting.objects.get(setting_name='H1stRange').configuration
    SelectionStage = AdminSetting.objects.get(setting_name='SelectionStage').configuration
    select_before_camp = AdminSetting.objects.get(setting_name='select_before_camp').configuration
    return render(request, 'updateData.html', {
        'J1stRange': J1stRange, 
        'H1stRange': H1stRange, 
        'SelectionStage': SelectionStage,
        'select_before_camp': select_before_camp
        })

def update_settings(request):
    if request.method == 'POST':
        J1stRange = request.POST['J1stRange']
        H1stRange = request.POST['H1stRange']
        SelectionStage = request.POST['SelectionStage']
        select_before_camp = request.POST['select_before_camp']
        print(select_before_camp)
        if (J1stRange == AdminSetting.objects.get(setting_name='J1stRange').configuration and
            H1stRange == AdminSetting.objects.get(setting_name='H1stRange').configuration and
            SelectionStage == AdminSetting.objects.get(setting_name='SelectionStage').configuration and
            select_before_camp == AdminSetting.objects.get(setting_name='select_before_camp').configuration):
            print(AdminSetting.objects.get(setting_name='select_before_camp').configuration)
            print('no change')
            return redirect('updateData')
        if (J1stRange and not J1stRange.isdigit()) or (J1stRange and not H1stRange.isdigit()):
            messages.error(request, '節次請輸入數字')
            return redirect('updateData')
        AdminSetting.objects.filter(setting_name='J1stRange').update(configuration=J1stRange) if J1stRange else None
        AdminSetting.objects.filter(setting_name='H1stRange').update(configuration=H1stRange) if H1stRange else None
        AdminSetting.objects.filter(setting_name='SelectionStage').update(configuration=SelectionStage)
        AdminSetting.objects.filter(setting_name='select_before_camp').update(configuration=select_before_camp)
        messages.success(request, '設定已更新')
        return redirect('updateData')
    return redirect('updateData')

def upload_zip(request): # 匯入資料壓縮檔（基本資料、志工大頭貼）
    valid_files = {'db_import.xlsx', 'pfp'}
    if request.method == 'POST' and request.FILES.get('uploadZip'):
        zip_file = request.FILES['uploadZip']
        if 'DBzip' not in zip_file.name:
            messages.error(request, '請按照檔名與格式要求上傳')
            return redirect('updateData')
            
        fs = FileSystemStorage()
        zip_filename = fs.save(zip_file.name, zip_file)
        zip_path = fs.path(zip_filename)
        extract_path = fs.location + '/' 
        
        # Create a specific folder for extraction
        # folder_name = 'DBzip'
        # extract_path = os.path.join(extract_folder, folder_name)


        # Extract the zip file into the folder
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # Create extraction path before extracting
            
            # Extract all files with correct encoding for Traditional Chinese
            for file_info in zip_ref.infolist():
                file_name = file_info.filename
                # Decode filename with cp437 and encode with utf-8 to handle Traditional Chinese
                try:
                    file_name = file_name.encode('cp437').decode('utf-8')
                except UnicodeDecodeError:
                    # Fallback if the encoding conversion fails
                    file_name = file_name.encode('cp437').decode('cp437')
                
                # Extract the file with the corrected filename
                source = zip_ref.open(file_info.filename)
                target_path = os.path.join(extract_path, file_name)
                
                # Create directories if needed
                if file_info.is_dir():
                    os.makedirs(target_path, exist_ok=True)
                    continue
                
                # Make sure directory exists
                os.makedirs(os.path.dirname(target_path), exist_ok=True)
                
                # Extract the file
                with open(target_path, 'wb') as target:
                    target.write(source.read())
            
            # Get a list of all extracted files
            extracted_files = {os.path.basename(file_info.filename) for file_info in zip_ref.infolist() 
                              if not file_info.is_dir()}
        extract_path = os.path.join(extract_path, 'DBzip')  # Ensure the extraction path is correct
        
        if not extracted_files.intersection(valid_files):
            try:
                shutil.rmtree(extract_path, ignore_errors=True)
            except (OSError, FileNotFoundError) as e:
                print(f"Error removing directory {extract_path}: {e}")
            
            fs.delete(zip_filename)
            messages.error(request, '請依照檔名與格式要求上傳')
            return redirect('updateData')

        if 'db_import.xlsx' in extracted_files:
            try:
                excel_path = os.path.join(extract_path, 'db_import.xlsx')
                print(f"Reading Excel file from: {excel_path}")
                dfs = pd.read_excel(excel_path, sheet_name=None)
                
                for sheet_name, df in dfs.items():
                    if sheet_name == 'student':
                        for _, row in df.iterrows():
                            student_instance, created = Student.objects.get_or_create(std_id=row['std_id'])
                            student_instance.std_id = row['std_id']
                            student_instance.std_name = row['std_name']
                            student_instance.team = row['team']
                            student_instance.satb = row['satb'].upper()
                            student_instance.j_or_h = row['j_or_h'].upper()
                            student_instance.std_tag = row.get('std_tag', '')
                            student_instance.save()
                    elif sheet_name == 'section':
                        for _, row in df.iterrows():
                            section_instance, created = Section.objects.get_or_create(section_id=row['section_id'])
                            section_instance.section_id = row['section_id']
                            section_instance.section_time = row['section_time']
                            section_instance.section_display = row['section_display']
                            section_instance.save()
                    elif sheet_name == 'volunteer':
                        for _, row in df.iterrows():
                            volunteer_instance, created = Volunteer.objects.get_or_create(volunteer_id=row['volunteer_id'])
                            volunteer_instance.volunteer_id = row['volunteer_id']
                            volunteer_instance.camp_name = row['camp_name']
                            volunteer_instance.save()
                    elif sheet_name == 'course':
                        for _, row in df.iterrows():
                            if row['course_type'] == 'js' or row['course_type'] == 'hs':
                                course_instance, created = SpecialCourse.objects.get_or_create(course_id=row['course_id'])
                                course_instance.std_limit = 99 if pd.isna(row['std_limit']) else row['std_limit']
                            else:
                                course_instance, created = Course.objects.get_or_create(course_id=row['course_id'])
                                course_instance.std_limit = 25 if pd.isna(row['std_limit']) else row['std_limit']
                            course_instance.course_id = row['course_id']
                            course_instance.course_name = row['course_name']
                            course_instance.course_info = row['course_info'] if pd.notna(row['course_info']) else '無課程資訊'
                            course_instance.course_type = row['course_type'].upper()
                            # Retrieve the Section instance
                            section_id = row.get('section_id')
                            course_instance.section_id = Section.objects.get(section_id=section_id)
                            course_instance.save()
            except Exception as e:
                messages.error(request, f'處理 Excel 檔案時發生錯誤: {str(e)}')
                print(f'Excel processing error: {e}')
                # Properly clean up any extracted files and the original zip file
                # try:
                #     # Remove the zip file
                #     if os.path.exists(zip_path):
                #         os.remove(zip_path)
                    
                #     # Remove the extracted directory and all its contents
                #     if os.path.exists(extract_path):
                #         shutil.rmtree(extract_path)
                #         print(f"Cleaned up extracted files from: {extract_path}")
                # except Exception as e:
                #     print(f"Error during cleanup: {e}")
                return redirect('updateData')
                
        # Cleanup: Remove the zip file but keep the extracted files if needed
        try:
            os.remove(zip_path)
            # # Remove extracted files if possible
            # try:
            #     shutil.rmtree(extract_path)
            #     print(f"Removed extracted files from: {extract_path}")
            # except (OSError, FileNotFoundError) as e:
            #     print(f"Error removing directory {extract_path}: {e}")
        except (OSError, FileNotFoundError) as e:
            print(f"Error removing file {zip_path}: {e}")
            
        messages.success(request, '資料匯入已完成')
        return redirect('updateData')
    return redirect('updateData')


def generate_xlsx(request): # 下載點名總表
    if request.method == 'POST':
        stage = request.POST['stage']
        selection_range = AdminSetting.objects.get(setting_name='J1stRange').configuration
        if stage == "1":
            sections = Section.objects.filter(section_id__lte=selection_range).order_by('section_id')
        else:
            sections = Section.objects.filter(section_id__gt=selection_range).order_by('section_id')
    # Create a workbook and add a worksheet for each section
    wb = openpyxl.Workbook()

    for section in sections:
        ws = wb.create_sheet(title=f"{section.section_display}")

        notation = [
            ("標示：", ""),
            ("國中部", "B3CEFA"),  # Junior
            ("高中部", "D2F1DA"),  # High
            ("CIT", ""),  # CIT
            ("", ""),
            ("solo：", ""),
            ("solo1", "FFFF00"),  # Yellow
            ("solo2", "FFA500"),  # Orange
            ("solo3", "D8BFD8"),  # Light Purple
        ]
        for row_num, (text, fill_color) in enumerate(notation, start=1):
            ws[f"A{row_num}"] = text
            if fill_color:
                ws[f"A{row_num}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Get all courses in the section
        courses = list(Course.objects.filter(section_id=section.section_id)) + list(SpecialCourse.objects.filter(section_id=section.section_id))

        col_num = 3
        for course in courses:
            # 新增、課程、教室、助教
            ws.merge_cells(start_row=3, start_column=col_num, end_row=3, end_column=col_num+4)
            ws.merge_cells(start_row=4, start_column=col_num, end_row=4, end_column=col_num+4)
            ws.merge_cells(start_row=5, start_column=col_num, end_row=5, end_column=col_num+4)
            ws[f"{get_column_letter(col_num)}3"] = course.course_name
            ws[f"{get_column_letter(col_num)}3"].font = Font(bold=True)
            ws[f"{get_column_letter(col_num)}4"] = "classroom"
            ws[f"{get_column_letter(col_num)}4"].font = Font(bold=True)
            ws[f"{get_column_letter(col_num)}5"] = "助教："
            ws[f"{get_column_letter(col_num)}5"].font = Font(bold=True)

            # Center align course_name, classroom, and TA
            for row in range(3, 6):
                ws[f"{get_column_letter(col_num)}{row}"].alignment = Alignment(horizontal="center")

            # Add table headers
            headers = ["人數", "學員", "SATB", "小隊", "點名格"]
            for i, header in enumerate(headers):
                ws[f"{get_column_letter(col_num + i)}6"] = header

            row_num = 7

            # Get all students in the course
            selections = SelectionResult.objects.filter(object_id=course.course_id).select_related('std').order_by('std__satb', 'std__j_or_h', 'std__team', 'std__std_name')
            for idx, selection in enumerate(selections, start=1):
                student = selection.std
                row = [
                    idx,
                    student.std_name,
                    student.satb,
                    student.team,
                    ""  # Placeholder for attendance checkbox
                ]
                for i, cell_value in enumerate(row):
                    ws[f"{get_column_letter(col_num + i)}{row_num}"] = cell_value

                # Apply background color based on j_or_h
                if student.std_tag == 'CIT':
                    fill_color = "FFFFFF"
                elif student.j_or_h == 'H':
                    fill_color = "D2F1DA"
                elif student.j_or_h == 'J':
                    fill_color = "B3CEFA"
                for col in range(col_num, col_num + 5):
                    ws[f"{get_column_letter(col)}{row_num}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Apply text color if student has tag "gr"
                if student.std_tag == "gr":
                    ws[f"{get_column_letter(col_num + 1)}{row_num}"].font = Font(color="FF0000")

                row_num += 1


            col_num += 6  # Move to the next set of columns for the next course

    # Remove the default sheet created by openpyxl
    wb.remove(wb['Sheet'])

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=selection_results.xlsx'
    wb.save(response)

    return response

def std_index(request): # 學生主頁（表單導覽頁）
    student = get_student(request.session.get('std_id'))

    formStatus = {}
    formStatus[1] = student.form1_completed
    formStatus[2] = student.form2_completed

    request.session['test'] = 0 #測試用

    select_before_camp = AdminSetting.objects.get(setting_name='select_before_camp').configuration
    SelectionStage = AdminSetting.objects.get(setting_name='SelectionStage').configuration
    context = {
        'formStatus': formStatus,
        'select_before_camp': select_before_camp,
        'SelectionStage': SelectionStage
    }
    if student is not None:
        print('student:', student)
    return render(request, 'std_index.html', {'student': student, 'context': context})
    

def select_form(request, form_stage): # 選課表單頁面
    if request.session.has_key('std_id'):
        std_id = request.session['std_id']
        team = request.session['team']
        student_instance = Student.objects.get(std_id=std_id)
        team_display = Student.TEAM_CHOICES[team-1][1]
        # current_form_stage = AdminSetting.objects.get(setting_name='SelectionStage').configuration
        form_type = student_instance.j_or_h + str(form_stage) #returning J1, J2, H1, H2
        form_display = dict(Student.FORM_DISPLAY)[form_type] #returning 第一次選課｜國中部, 第二次選課｜國中部, 第一次選課｜高中部, 第二次選課｜高中部


        student = {
            'std_id': std_id,
            'team_display': team_display,
            'form_display': form_display,
            'std_name': student_instance.std_name
        }
        request.session['team_display'] = team_display
        request.session['form_display'] = form_display
        request.session['form_stage'] = form_stage
        request.session['form_type'] = form_type
        print('form_stage:', form_stage)
        


        return render(request, 'select_form.html', {
            'student': student, 
            'team_display': team_display, 
            'form_display': form_display, 
            })
    else:
        return redirect('/stdLogin')

def get_courses(request): # 從資料庫抓課程回傳到前端
    std_id = request.session['std_id']
    student_instance = Student.objects.get(std_id=std_id)
    selection_range = AdminSetting.objects.get(setting_name=student_instance.j_or_h + '1stRange').configuration
    form_stage = request.session['form_stage']
    print('form_stage in get_courses', form_stage)
    if form_stage == 1:
        section_instances = Section.objects.filter(section_id__lte=selection_range)
        course_instances = Course.objects.filter(section_id__lte=selection_range, course_type__in=[student_instance.j_or_h, 'M', 'NA'])
        sp_course_instances = SpecialCourse.objects.filter(section_id__lte=selection_range, course_type=student_instance.j_or_h+'S') # 'JS'國中部課程, 'HS'高中部課程
        print(sp_course_instances)
    else:
        section_instances = Section.objects.filter(section_id__gt=selection_range)
        course_instances = Course.objects.filter(section_id__gt=selection_range, course_type__in=[student_instance.j_or_h, 'M', 'NA'])
        sp_course_instances = SpecialCourse.objects.filter(section_id__gt=selection_range, course_type=student_instance.j_or_h+'S') # 'JS'國中部課程, 'HS'高中部課程

    # 課程類別 course_type:
    # 'J' 國中限定
    # 'H' 高中限定
    # 'M' 國高中混合
    # 'JS' 國中部課程
    # 'HS' 高中部課程
    # 'NA' 連堂第二節
    
    sections_with_courses = [] # list中插入dict用來分開儲存每一個節次對應的所有課程，以及課程數量，用來顯示志願選項數量
    for section in section_instances:
        courses_in_section = []
        sp_courses_in_section = sp_course_instances.filter(section_id=section.section_id)
        if sp_courses_in_section.exists():
            for course in sp_courses_in_section:
                if '_' in course.course_name:
                    if '、' in course.course_name:
                        teachers =  course.course_name.split('_')[1].split('、') 
                    else:
                        teachers = [course.course_name.split('_')[1]]
                else: teachers = []
                courses_in_section.append({
                    'course_id': course.course_id,
                    'course_name': course.course_name,
                    'course_info': course.course_info,
                    'course_type': course.course_type,
                    'teachers': teachers
                })
                print('SP_course:', course.course_name)
        else:
            for course in course_instances.filter(section_id=section.section_id):
                if '_' in course.course_name:
                    if '、' in course.course_name:
                        teachers =  course.course_name.split('_')[1].split('、') 
                    else:
                        teachers = [course.course_name.split('_')[1]]
                else: teachers = []
                courses_in_section.append({
                    'course_id': course.course_id,
                    'course_name': course.course_name,
                    'course_info': course.course_info,
                    'course_type': course.course_type,
                    'teachers': teachers
                })
        num_courses = len(courses_in_section)
        sections_with_courses.append({
            'section_id': section.section_id,
            'section_time': section.section_time,
            'section_display': section.section_display,
            'courses': courses_in_section,
            'num_courses': num_courses
        })

    return JsonResponse({
        'sections_with_courses': sections_with_courses,
        'saved_selections': request.session.get('selections', []),
        'test' : request.session['test']
    })

def double_check(request): # merge with confirm
    if request.method == 'POST':
        selections = request.POST.getlist('priority')
        std_id = request.session['std_id']


        # Temporarily save the user's choices in the session
        request.session['selections'] = selections

        # print('selections:', selections)
        return redirect('confirm')

    return redirect('select_form')

def confirm(request): # 確認結果頁面
    request.session['test'] = 1

    team_display = request.session['team_display']
    form_display = request.session['form_display']
    selections = request.session.get('selections', [])
    std_id = request.session['std_id']
    student = Student.objects.get(std_id=std_id)

    # Process selections to generate the table data
    table_data = {}
    priorities = set()

    for selection in selections:
        priority, course_id, std_id, section_id = selection.split('-')
        if SpecialCourse.objects.filter(course_id=course_id).exists():
            course = SpecialCourse.objects.get(course_id=course_id)
        else:
            course = Course.objects.get(course_id=course_id)
        section = Section.objects.get(section_id=section_id)
        if section.section_display not in table_data:
            table_data[section.section_display] = {}
        table_data[section.section_display][priority] = course.course_name
        priorities.add(priority)


    # Sort priorities
    sorted_priorities = sorted(priorities, key=int)

    return render(request, 'confirm.html', {
        'student': student,
        'table_data': table_data,
        'sorted_priorities': sorted_priorities,
        'team_display': team_display,
        'form_display': form_display,
        'form_stage': request.session['form_stage'],
        'test' : request.session['test']
    })

def submit_form(request): # 提交志願表單
    request.session['test'] = 0

    selections = request.session.get('selections', [])
    std_id = request.session['std_id']  # Assuming std_id is used as the username
    form_type = request.session['form_type']

    # Process and save the selections to the database
    for selection in selections:
        priority, course_id, std_id, section_id = selection.split('-')

        if SpecialCourse.objects.filter(course_id=course_id).exists():
            course_instance = SpecialCourse.objects.get(course_id=course_id)
        else:
            course_instance = Course.objects.get(course_id=course_id)

        selection_instance = Selection.objects.create(
            priority=priority,
            std=Student.objects.get(std_id=std_id),
            course_id=course_instance.course_id,
            section=Section.objects.get(section_id=section_id),
            form_type=form_type
        )

    # Clear the selections from the session
    del request.session['selections']
    return redirect('success')  # Redirect to a success page or another appropriate page

def success(request): # 選課成功
    std_id = request.session['std_id']
    student = Student.objects.get(std_id=std_id)
    form_stage = request.session['form_stage']


    if form_stage == '1':
        student.form1_completed = True
        print('form1 completed')
    elif form_stage == '2':
        student.form2_completed = True
        print('form2 completed')
    # student.save() 測試先關掉
    return render(request, 'success.html')

def stdLogout(request):
    std_logout(request)
    request.session.clear()
    return redirect('/stdLogin')

def logout(request):
    std_logout(request)
    return redirect('index')



def process_excel_form(request): # 處理志願表單
    if request.method == 'POST' and request.FILES.get('upload_excel_form'):
        excel_file = request.FILES['upload_excel_form']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        course_not_found = False

        # Read the Excel file
        df = pd.read_excel(file_path)

        # Standardize column names
        df.columns = df.columns.str.strip().str.lower()

        # Check if 'std_id' column exists
        if 'std_id' not in df.columns:
            # fs.delete(filename)
            messages.error(request, 'std_id 欄位不存在')
            return redirect('updateData')

        # Process the data
        unfound_courses = []
        dfs = pd.read_excel(file_path, sheet_name=None)
        for sheet_name, df in dfs.items():
            form_type = sheet_name if sheet_name in ['J1', 'J2', 'H1', 'H2'] else None
            if not form_type:
                messages.error(request, '工作表名稱必須是 J1, J2, H1, H2')
                return redirect('updateData')
            for index, row in df.iterrows():
                print(row)
                std_id = row['std_id']
                try:
                    student = Student.objects.get(std_id=std_id)
                except Student.DoesNotExist:
                    fs.delete(filename)
                    messages.error(request, f'Student with std_id {std_id} does not exist')
                    return redirect('updateData')

                for col in df.columns:
                    if col != 'std_id':
                        section_id, priority = col.split('_')
                        course_name = row[col]
                        course_instance = Course.objects.filter(section_id=section_id)
                        special_course_instance = SpecialCourse.objects.filter(section_id=section_id)
                        if course_instance.filter(course_name=course_name).exists():
                            print('course_name:', course_name)
                            print('section_id:', section_id)
                            print('priority:', priority)
                            course = course_instance.get(course_name=course_name)
                        elif special_course_instance.filter(course_name=course_name).exists():
                            course = special_course_instance.get(course_name=course_name)
                        else:
                            if course_name not in unfound_courses:
                                unfound_courses.append(course_name)
                            course_not_found = True
                            continue

                        section = Section.objects.get(section_id=section_id)

                        # Save the selection to the database
                        if not course_not_found:
                            Selection.objects.create(
                                priority=priority,
                                std=student,
                                course_id=course.course_id,
                                section=section,
                                form_type=form_type
                            )


        # Clean up the uploaded file
        if unfound_courses:
            fs.delete(filename)
            messages.error(request, '課程無法對應資料庫：' + ', '.join(unfound_courses))
            return redirect('updateData')

        fs.delete(filename)
        messages.success(request, '資料匯入已完成')
        return redirect('updateData')

    return render(request, 'updateData.html')

def truncate_table(model): # 清空資料表
    with connection.cursor() as cursor:
        table_name = model._meta.db_table
        cursor.execute(f'TRUNCATE TABLE `{table_name}`')


def truncate_data(request): # 選課組後台-清空選課資料 / 志願結果表格
    if request.method == 'POST':
        model_name = request.POST.get('model')
        if model_name == 'Selection':
            truncate_table(Selection)
            messages.success(request, '學員選課資料已刪除')
        elif model_name == 'SelectionResult':
            truncate_table(SelectionResult)
            messages.success(request, '志願結果已刪除')
        elif model_name == 'All':
            truncate_table(Selection)
            truncate_table(SelectionResult)
            Student.objects.all().delete()
            Section.objects.all().delete()
            Course.objects.all().delete()
            SpecialCourse.objects.all().delete()
            Volunteer.objects.all().delete()
            media_storage = FileSystemStorage().location
            # Remove all files in the media storage directory
            for filename in os.listdir(media_storage):
                file_path = os.path.join(media_storage, filename)
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f"Error removing file {file_path}: {e}")
            messages.success(request, '所有資料已經除')
        else:
            messages.error(request, '無效的模型名稱')

    return redirect('updateData')



def process_selection_results(request): # 處理志願結果
    if request.method == 'POST':
        processing_stage = request.POST['processing_stage']
        # 從按鈕判斷處理第一階或第二階選課
        if processing_stage == '1':
            sections = Section.objects.filter(section_id__lte=AdminSetting.objects.get(setting_name='J1stRange').configuration) # 1~6
        elif processing_stage == '2': 
            sections = Section.objects.filter(section_id__gt=AdminSetting.objects.get(setting_name='J1stRange').configuration) # 7~12
        
        # Process each section independently
        for section in sections:
            print(f"Processing section {section.section_id}")
            
            # Get all courses in this section
            courses_in_section = list(Course.objects.filter(section_id=section.section_id))
            special_courses_in_section = list(SpecialCourse.objects.filter(section_id=section.section_id))
            
            # Combine regular and special courses for processing
            all_courses_in_section = courses_in_section + special_courses_in_section
            
            # Dictionary to track assigned students to courses at each priority level
            course_assignments = {}
            
            # Dictionary to track course limits
            course_limits = {}
            
            # Dictionary to track course types
            course_types = {}
            
            # Track pre-allocated students to protect them from de-allocation
            protected_students = set()
            
            # Initialize course data
            for course in courses_in_section:
                course_assignments[course.course_id] = []
                course_limits[course.course_id] = course.std_limit
                course_types[course.course_id] = course.course_type
                
            for special_course in special_courses_in_section:
                # Use a special prefix to distinguish SpecialCourse IDs
                special_key = f"S{special_course.course_id}"
                course_assignments[special_key] = []
                course_limits[special_key] = special_course.std_limit
                course_types[special_key] = special_course.course_type
            
            # Get all selections for this section
            all_selections = Selection.objects.filter(section=section)
            
            # Group selections by student
            student_selections = {}
            for selection in all_selections:
                std_id = selection.std_id
                if std_id not in student_selections:
                    student_selections[std_id] = []
                
                # Parse course_id to determine if it's a regular or special course
                course_id_str = selection.course_id
                
                # Check if this is a special course ID by trying to find it in SpecialCourse
                try:
                    special_course = SpecialCourse.objects.get(course_id=int(course_id_str))
                    course_key = f"S{special_course.course_id}"
                except (SpecialCourse.DoesNotExist, ValueError):
                    # Not a special course, so use the regular course ID
                    course_key = int(course_id_str)
                
                student_selections[std_id].append({
                    'course_key': course_key,
                    'priority': selection.priority
                })
            
            # Sort each student's selections by priority
            for std_id in student_selections:
                student_selections[std_id].sort(key=lambda x: int(x['priority']))
            
            # Track assigned and unassigned students
            assigned_students = set()
            unassigned_students = set(student_selections.keys())
            
            # Clear previous results for this section
            SelectionResult.objects.filter(section=section).delete()
            
            # =================== PRE-ALLOCATION RULES ===================
            
            # 1. Sign Language Course Assignment (光仁學生先全部指定到手語課)
            sl_courses = []
            for course in courses_in_section:
                if '手語課' in course.course_name:
                    sl_courses.append(course)
            
            for special_course in special_courses_in_section:
                if '手語課' in special_course.course_name:
                    sl_courses.append(special_course)
            
            sl_selections = []
            
            if sl_courses:
                # Get all students with std_tag='gr'
                gr_students = Student.objects.filter(std_tag='gr')
                
                for gr_student in gr_students:
                    if gr_student.std_id in unassigned_students:
                        # Assign to the first available sign language course
                        course = sl_courses[0]  # Use the first sign language course
                        
                        # Determine course key and content type
                        if isinstance(course, SpecialCourse):
                            course_key = f"S{course.course_id}"
                            content_type = ContentType.objects.get_for_model(SpecialCourse)
                        else:
                            course_key = course.course_id
                            content_type = ContentType.objects.get_for_model(Course)
                        
                        # Add to course assignments
                        course_assignments[course_key].append(gr_student.std_id)
                        
                        # Mark as assigned and protected
                        assigned_students.add(gr_student.std_id)
                        protected_students.add(gr_student.std_id)
                        unassigned_students.remove(gr_student.std_id)
                        
                        # Create SelectionResult object
                        sl_selections.append(SelectionResult(
                            std=gr_student,
                            section=section,
                            content_type=content_type,
                            object_id=course.course_id,
                            form_type=processing_stage
                        ))
                
                # Save all sign language assignments at once
                SelectionResult.objects.bulk_create(sl_selections)
                print(f"Pre-allocated {len(sl_selections)} GR students to sign language courses")
            
            # 2. Consecutive Course Assignment (處理連堂的第二堂課)
            na_selections = []
            
            for course in courses_in_section:
                if course.course_type == 'NA':
                    try:
                        # Find the first part of this course
                        previous_course_id = course.course_id - 1
                        previous_course = Course.objects.get(course_id=previous_course_id)
                        previous_results = previous_course.selection_results.all()
                        
                        # # Find all students assigned to the previous course
                        # previous_results = SelectionResult.objects.filter(
                        #     section=section,
                        #     content_type=ContentType.objects.get_for_model(Course),
                        #     object_id=previous_course_id
                        # )
                        
                        for result in previous_results:
                            student = result.std
                            
                            # Only process if student is not yet assigned for this section
                            if student.std_id in unassigned_students:
                                # Add to course assignments
                                course_assignments[course.course_id].append(student.std_id)
                                
                                # Mark as assigned and protected
                                assigned_students.add(student.std_id)
                                protected_students.add(student.std_id)
                                unassigned_students.remove(student.std_id)
                                
                                # Create SelectionResult object
                                na_selections.append(SelectionResult(
                                    std=student,
                                    section=section,
                                    content_type=ContentType.objects.get_for_model(Course),
                                    object_id=course.course_id,
                                    form_type=processing_stage
                                ))
                    except Course.DoesNotExist:
                        print(f"Previous course for NA course {course.course_id} not found")
            
            # Save all consecutive course assignments at once
            SelectionResult.objects.bulk_create(na_selections)
            print(f"Pre-allocated {len(na_selections)} students to consecutive NA courses")
            
            # =================== MAIN ALLOCATION ALGORITHM ===================
            
            # Maximum priority to consider
            max_priority = max([len(selections) for selections in student_selections.values()]) if student_selections else 0
            
            # Allocation algorithm
            for priority_level in range(max_priority):
                print(f"Processing priority level {priority_level + 1}")
                
                # Attempt assignment for current priority level
                for std_id in list(unassigned_students):
                    selections = student_selections[std_id]
                    
                    # Skip if student doesn't have this priority level
                    if priority_level >= len(selections):
                        continue
                    
                    # Get course for current priority
                    course_key = selections[priority_level]['course_key']
                    
                    # Add student to course assignments
                    course_assignments[course_key].append(std_id)
                    assigned_students.add(std_id)
                    unassigned_students.remove(std_id)
                
                # Handle over-subscription
                for course_key, assigned in course_assignments.items():
                    # Skip if course type is 'NA' (unlimited)
                    # if course_types.get(course_key) == 'NA':
                    #     continue
                    
                    limit = course_limits.get(course_key, 0)
                    
                    # Filter out protected students for de-allocation consideration
                    unprotected_students = [std_id for std_id in assigned if std_id not in protected_students]
                    
                    # Calculate excess based on unprotected students that can be removed
                    total_assigned = len(assigned)
                    protected_count = total_assigned - len(unprotected_students)
                    
                    # If we have more students than the limit, and there are unprotected students to remove
                    excess = total_assigned - limit
                    if excess > 0 and unprotected_students:
                        # Limit the excess to the number of unprotected students
                        excess = min(excess, len(unprotected_students))
                        
                        # Randomly select students to unassign from unprotected students only
                        import random
                        students_to_unassign = random.sample(unprotected_students, excess)
                        
                        # Update assignments
                        for std_id in students_to_unassign:
                            assigned.remove(std_id)
                            assigned_students.remove(std_id)
                            unassigned_students.add(std_id)
                
                # Break if all students are assigned
                if not unassigned_students:
                    print("All students assigned successfully")
                    break
            
            # Save results to database for regular allocation
            selection_results = []
            
            for course_key, assigned_std_ids in course_assignments.items():
                for std_id in assigned_std_ids:
                    student = Student.objects.get(pk=std_id)
                    
                    # Skip if this is a pre-allocated student that was already saved
                    if std_id in protected_students and (
                        student.std_tag == 'gr' or 
                        any(course.course_type == 'NA' for course in courses_in_section if course.course_id == course_key)
                    ):
                        continue
                    
                    # Determine if this is a special course or regular course
                    if isinstance(course_key, str) and course_key.startswith('S'):
                        # Special course
                        special_course_id = int(course_key[1:])
                        special_course = SpecialCourse.objects.get(course_id=special_course_id)
                        
                        # Create the content type for SpecialCourse
                        content_type = ContentType.objects.get_for_model(SpecialCourse)
                        
                        selection_results.append(SelectionResult(
                            std=student,
                            section=section,
                            content_type=content_type,
                            object_id=special_course_id,
                            form_type=processing_stage
                        ))
                    else:
                        # Regular course
                        course = Course.objects.get(course_id=course_key)
                        
                        # Create the content type for Course
                        content_type = ContentType.objects.get_for_model(Course)
                        
                        selection_results.append(SelectionResult(
                            std=student,
                            section=section,
                            content_type=content_type,
                            object_id=course.course_id,
                            form_type=processing_stage
                        ))
            
            # Save all regular allocation results at once
            SelectionResult.objects.bulk_create(selection_results)
            
            # Update assignment status for assigned students
            Student.objects.filter(std_id__in=list(assigned_students)).update(assigned=True)
            
            # Handle unassigned students (if any)
            if unassigned_students:
                print(f"Warning: {len(unassigned_students)} students could not be assigned to any course")
        
        messages.success(request, '志願分發完成')
        return redirect('result')
    
    return render(request, 'process_selection.html')

def print_results_table(request): # 選課組後台-列印志願結果
    # Gather data
    if request.method == 'POST':
        stage = request.POST['stage']
        selection_range = AdminSetting.objects.get(setting_name='J1stRange').configuration
        if stage == "1":
            sections = Section.objects.filter(section_id__lte=selection_range).order_by('section_id')
        else:
            sections = Section.objects.filter(section_id__gt=selection_range).order_by('section_id')
        students = Student.objects.all().order_by('team', 'std_id')
        selection_results = SelectionResult.objects.filter(section__in=sections)
        
        # Organize data
        Junior_data = {}
        High_data = {}
        for student in students.filter(j_or_h='J'):
            team = student.team
            team_display = Student.TEAM_CHOICES[team-1][1]
            if team not in Junior_data:
                Junior_data[team] = {'students': [], 'results': {}, 'team_display': team_display}
            Junior_data[team]['students'].append(student)
            for section in sections:
                if section.section_id not in Junior_data[team]['results']:
                    Junior_data[team]['results'][section.section_id] = {}
                result = selection_results.filter(std=student, section=section).first()
                if result:
                    Junior_data[team]['results'][section.section_id][student.std_id] = result.course.course_name
                else:
                    print('No result found for student', student.std_id, 'in section', section.section_id)
                    Junior_data[team]['results'][section.section_id][student.std_id] = ''
        for student in students.filter(j_or_h='H'):
            team = student.team
            team_display = Student.TEAM_CHOICES[team-1][1]
            if team not in High_data:
                High_data[team] = {'students': [], 'results': {}, 'team_display': team_display}
            High_data[team]['students'].append(student)
            for section in sections:
                if section.section_id not in High_data[team]['results']:
                    High_data[team]['results'][section.section_id] = {}
                result = selection_results.filter(std=student, section=section).first()
                if result:
                    High_data[team]['results'][section.section_id][student.std_id] = result.course.course_name
                else:
                    print('No result found for student', student.std_id, 'in section', section.section_id)
                    High_data[team]['results'][section.section_id][student.std_id] = ''

        return render(request, 'print_results_table.html', {
            'sections': sections,
            'Junior_data': Junior_data,
            'High_data': High_data,
        })



# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠟⠛⠋⠉⠉⠉⠉⠉⠉⠉⠉⠉⠉⠉⠙⠛⠿⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠉⠛⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠟⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠛⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠛⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠃⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣀⣀⣀⠀⣀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣤⣴⡄⠸⣿⣿⣇⠸⣿⣿⡏⢰⣿⠇⣸⣿⡿⢀⣿⡆⢰⡆⢠⣄⠀⠀⠀⠙⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣀⠐⢾⣆⠘⠿⠛⠀⠙⡋⠉⠀⠉⠉⠀⣚⣛⣀⣙⣛⡃⠘⠿⠀⠿⠃⣼⡏⢰⣷⠀⡀⢹⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⣶⡈⢻⣧⠈⠉⣀⣂⣭⣥⣶⣶⣿⣿⣿⣿⣶⣦⣭⣭⣭⣭⣭⣭⣭⡃⣶⣦⣬⡀⠻⠏⢀⡇⠈⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⢤⣶⣄⠹⠿⠀⣠⣶⣿⣿⣿⣿⠿⠿⠿⠿⠿⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠿⠿⢿⣿⣿⣷⣄⡈⢁⠀⢸⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⠀⢀⠐⢷⣄⡉⠛⣁⣴⣾⣿⣿⣋⣥⣶⣶⣾⣿⣿⣿⣿⡖⢿⣿⣿⣿⣿⣿⣿⣯⢴⣿⣿⣿⣶⣌⠻⣿⣿⣧⠈⠀⢸⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀⠀⣄⠙⢷⡄⠋⣠⣾⣿⣿⣿⣿⣿⣿⠛⣛⣛⣋⣩⣿⣿⣿⣿⣆⢹⣿⣿⣿⣿⣿⣿⠀⣿⣿⣿⠿⠿⣿⣿⣿⣿⠀⠁⣾⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⢴⣦⡙⠳⠄⢀⣾⣿⣿⣿⣿⣿⣿⢿⣿⣿⣿⠿⠿⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣤⣿⣿⣿⣿⣷⣶⣬⣙⣻⠇⣸⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⣶⣄⠙⢿⡆⢠⣿⣿⠿⣿⣿⣿⣿⣃⣀⣨⣇⠐⢶⠀⠀⠀⠀⣠⢸⣿⣿⣿⣿⣿⣿⣿⣿⣿⠋⣉⡉⠛⠛⠛⣿⡿⢯⠀⣿⣿⣿⣿⣿⣿⣿⢿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣇⠀⣦⠘⢿⣶⠄⢠⣾⣿⣿⡎⠘⠛⣿⣿⣿⣿⣿⡏⣁⣀⢀⣀⣀⠀⠀⠈⣿⣿⣿⣿⣿⣿⣿⣿⣏⠠⡉⠁⠀⠀⠀⣿⣄⠛⠀⢿⣿⣿⣿⣿⣿⡏⣼⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣆⠹⣷⣄⠙⢠⣿⣿⣿⣿⡧⠈⡀⣻⣿⣿⣿⣿⣧⣬⣍⣀⠀⠀⠈⢀⣼⣿⣿⣿⣿⣿⣿⣿⠈⣿⣦⣀⠈⠁⠀⠰⠛⣿⣿⡇⠈⣿⣿⣿⣿⣿⠇⣾⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣆⠹⣿⡇⢸⣿⣿⣿⣿⡏⠀⠓⢌⠼⢻⣿⣿⣿⣿⣿⣿⣷⣶⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⡄⠹⣿⣿⣦⣼⣿⣿⣿⣿⣿⣥⠀⣿⣿⣿⣿⠟⠁⣰⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣦⡈⠃⣸⣿⣿⣿⣿⡇⢸⡓⣎⡇⠧⠙⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣄⠙⢿⣿⣿⣿⡿⢷⡿⣏⠢⠀⣿⣿⡏⠌⣴⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠦⠘⣿⣿⣿⣿⣇⠈⠼⡐⢛⢀⠺⣿⡙⢿⣿⣿⣿⣿⣿⣿⠟⢛⣉⣿⣿⣿⣿⣿⣿⣿⣷⠀⣿⣿⣿⣿⣮⡀⢠⠅⣰⣿⣿⠘⣼⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣇⡀⢸⣿⣿⣿⣿⡆⠨⢼⣆⡻⣦⠹⣛⣧⣿⣿⣿⣿⡟⢿⡄⠹⣿⣿⡿⠿⢿⣿⣿⣿⠁⣼⣿⣿⣏⡟⡥⠀⠀⢠⣿⣿⡆⢿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠁⢸⣿⣿⣿⣿⣷⡄⠸⣡⠯⠛⣖⡆⣽⣿⣍⣿⡟⢡⡾⢿⠶⣶⣾⣗⣿⡿⣛⣶⣤⣼⡁⣿⣟⣋⠌⡶⠄⢠⣿⣿⡟⠈⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠃⢸⣿⣿⣿⣿⣿⠹⡄⠁⠜⣷⣾⢻⣭⡀⣶⡟⣰⣿⡉⣟⢋⣤⣼⡮⠂⢖⡋⠓⠺⠿⣇⢹⣷⡃⡅⠃⣰⣿⣿⣿⡗⢀⢸⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣏⠀⢸⣿⣿⣿⣿⣿⡇⣷⢤⡄⠉⠻⢃⢥⡛⣾⣤⣿⣿⣁⣌⣙⡛⠻⠿⠿⠿⠿⠛⠐⠂⢌⡀⢿⡩⠄⣴⣿⣿⣿⣿⡿⢸⠘⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠛⠛⠋⠁⠀⣾⣿⣿⣿⣿⣿⠀⣗⣿⣿⣧⣄⠀⠁⢳⣬⠘⢻⣿⣿⢿⣿⣿⣏⡙⣉⢙⣛⣁⡒⣌⡀⢍⠫⠉⠀⣿⣿⣿⣿⣿⠇⠜⢠⣿⣿⣿⣿⣿
# ⣿⣿⣿⣿⣿⠿⠟⠉⠁⠀⠀⠀⠀⠀⠀⣼⣿⣿⣿⣿⣿⠧⢴⠿⡼⢞⣦⣙⣏⣦⠀⠻⢳⣖⣾⠻⣯⣝⢳⣩⡿⡿⢛⣻⢟⠿⡏⣛⡆⠈⠂⢔⠻⣿⠟⡫⠔⣋⣴⣿⣿⣿⣿⣿⣿
# ⣿⣿⣿⠏⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣿⣿⣿⣿⣿⣿⣇⠲⠭⢶⣵⠖⣙⣬⡿⣧⣆⠀⠘⢾⢉⣿⣿⣆⣿⢋⣽⣿⣴⡏⣭⣧⠛⠃⢴⣿⣶⣍⢀⣤⣶⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⣿⠟⠉⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⠿⠛⢉⣶⠽⣭⣾⣿⡷⢽⡻⣿⣷⣇⢄⢈⡙⠛⠇⠾⠿⠷⠻⠿⠟⠋⠁⠀⠀⠀⠙⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠚⠿⣿⣿⣿⣿⠁⣴⣿⣿⣿⣿⣿⣦⣿⣎⣢⡝⣶⣦⣹⣮⣠⠸⣟⣶⣶⣷⡦⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠻⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠁⢾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣿⣷⣝⢻⣿⣧⣘⠻⣿⣿⣿⣦⣀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⢿⣿⣿⣿⣿⣿⣿⣿⣿
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠙⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣧⣿⣾⣿⣷⣿⣾⡿⠟⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠉⠉⠉⠙⠛⠿⣿⣿
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢉⣙⡛⠋⠩⠍⠻⠿⢿⡿⠟⣛⡛⠿⠛⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠻⡧⣌⢉⣁⣀⢀⣀⣀⢒⡶⠖⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀

# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠺⠽⠾⢟⣓⢦⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠛⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡀⠀⠀⣀⣀⣀⢤⣀⣀⣀⣀⣀⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣀⢤⣴⣲⣶⢶⠺⠾⠟⡛⢋⠛⠉⡉⠎⠉⡉⠙⠒⠛⠿⢾⡿⢱⣶⣦⣀⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⢴⣾⡿⣟⣻⠍⡉⢀⢁⠀⠂⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠀⠀⠉⠒⠈⠉⠚⠳⠻⠷⢲⡤⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⠤⣶⣿⠿⠉⠁⠆⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠈⡙⢯⣷⡦⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣰⣾⠿⠯⠅⠠⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠻⡷⢶⣄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣀⣴⣿⠻⠙⠀⠂⠀⠀⠀⢢⣴⣶⣶⣶⣶⣿⣿⣿⣿⣿⣿⣿⣶⣦⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣰⣶⣿⣿⣿⣿⣿⣿⣶⣶⣶⣶⣄⠈⠛⢿⡢⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣴⠿⠛⡉⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣟⣿⣯⣿⣾⣯⣿⣽⣿⣿⡿⠄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠿⣿⣿⣯⣷⣿⣷⣿⣯⣿⣿⣿⣿⢠⠈⠀⠹⢿⣧⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⡦⣶⣢⣦⡀⣀⢀⣶⠿⡋⠂⠁⠀⠀⠀⠀⠀⠀⠀⠐⡈⢿⣿⣿⡿⠿⠟⠛⠛⠉⠉⠉⠀⠀⠐⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠀⠀⠀⠉⠉⡙⠛⠛⠿⠿⣿⣿⠏⠀⠀⠀⠀⠈⣹⢳⣦⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⢀⣀⡀⠀⠀⠀⢀⡴⠏⠁⠀⠀⢻⣿⣷⣿⡟⠈⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⠢⠉⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠁⠀⠀⠒⠲⠤⠊⠀⠀⠀⠀⠀⠘⠮⣱⣷⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⢀⠞⠉⠀⠩⢲⣤⡶⠉⠀⠀⠀⠀⠀⢀⣿⣿⢃⠈⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠐⢼⣷⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⢠⠫⠀⠀⠀⠀⠈⢿⠀⠀⡀⠀⠀⠀⣠⣾⡿⢈⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠐⢀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⣝⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⡼⢯⠀⠀⠀⠀⢀⣿⠀⠀⠈⠓⢦⡿⠿⠏⠉⠉⠙⠲⣄⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠳⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣈⢄⡈⠄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⠞⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⡟⣸⠐⠤⠤⠴⠾⣳⠀⠀⠀⡴⣿⠃⠀⠀⠀⡀⠀⠀⠀⠙⠲⣄⠀⠀⠀⠀⠀⠀⠀⣀⡀⠄⠄⣀⣠⣤⣤⣤⣤⣦⣤⣤⣄⣀⡀⠀⠀⠀⠀⠘⢦⠀⠀⠀⠀⠀⠀⠀⠀⠀⡀⠀⢀⣀⣤⣴⣶⣶⣶⣶⣶⣶⣦⣤⣄⣀⠀⠀⠀⠀⠀⠀⠻⢽⣤⡤⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⣽⢺⡀⠀⠀⠀⠠⣿⠀⠓⠲⡏⠁⣄⠀⠀⢀⣧⡀⠀⠀⠀⠀⣸⡇⠀⠀⠀⣀⣀⣬⣷⣶⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣦⣄⣀⠀⠅⠀⠀⣀⣠⠤⢂⣀⣤⣶⣿⣿⣿⣿⣿⡿⠿⠿⠿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣶⣤⣤⣴⣿⣿⣿⣸⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⢼⡆⠀⣀⣀⣀⡼⢯⡀⠀⢘⡇⠀⣼⣷⣶⣿⣿⣷⣕⣀⣀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠿⠛⠛⠉⠉⠉⠉⠉⠉⠛⠙⠛⠿⣿⣿⣿⣿⣿⣷⣶⣤⣤⣶⣾⣿⣿⣿⣿⣿⠿⠛⠉⠀⠀⠀⠀⠀⠨⠀⠀⠉⠛⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⢾⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠸⣷⡀⠀⠀⠀⠀⢸⣷⠀⣽⡇⣠⣿⠁⠁⠉⢉⠛⣿⡿⠋⠛⢿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠃⠀⢀⢠⣀⣀⣀⡀⠄⠀⠀⢀⣘⣶⣀⣀⠙⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠫⣀⣤⣤⣴⣦⣤⣕⡀⣄⣤⣴⣿⣷⣤⡀⢸⣿⣿⣿⣿⡿⠛⠉⠹⣿⢸⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⢀⣿⡇⠀⠀⠀⠀⠀⣿⣿⣿⡷⣿⡃⢁⠀⠀⢀⣸⠁⠀⠀⠀⢸⡇⠈⢻⣿⣿⣿⣿⡏⠀⢀⡼⣿⣿⣿⣿⣿⣿⣷⣶⣿⣿⣿⣿⣿⣿⡄⠐⢻⣿⣿⣿⣿⣿⣿⣿⣿⡏⠀⡸⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡷⠀⢹⣿⣿⣿⠇⡆⠀⠐⢿⣏⡆⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⡾⠭⠁⠀⠀⠀⠀⠀⢾⣿⡿⣽⡿⠙⠲⠤⣰⠏⠚⢦⡀⠀⠀⣼⠇⠀⠀⣿⣿⣿⣿⡇⠀⠈⡇⢻⣿⣿⣿⣿⣿⣿⠻⣿⣿⣿⣿⣿⣿⠃⠀⠨⣿⣿⣿⡷⣿⣿⣿⣿⠀⠀⠁⠘⢿⣿⣿⣿⣿⣿⠃⠻⣿⣿⣿⣿⡿⠃⠀⣸⣿⣿⣿⢠⠃⠀⠈⡼⣿⢸⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⢀⣼⡏⠁⠀⠀⠀⠀⠀⠀⣿⣿⠟⠋⠀⠀⠀⠀⠈⢣⡀⠀⡉⠆⣠⠟⠀⠀⠀⢿⣿⣿⣿⣇⠀⠀⣀⡈⡙⠿⠿⠿⠟⠁⠀⠉⠛⢿⣟⣿⠥⠀⠀⠨⣿⣿⣿⡇⣿⣿⣿⣿⠀⠀⢾⢤⡬⡭⢿⣛⣉⣀⣀⣀⠬⠽⢽⡿⠒⠀⢠⣿⣿⣿⠃⠎⠀⠀⠀⢻⣿⠃⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⢸⣿⠇⠀⠀⠀⠀⠀⠀⢠⣿⠉⠁⠀⠀⠀⠀⠀⠀⠀⠃⠀⡶⠋⠁⠀⠀⠀⠀⠈⣿⣿⣿⣿⡄⠀⠉⠙⢾⡍⣏⣉⣉⣩⠭⢭⢩⠥⣟⠁⠀⡄⠀⣸⣿⣿⡿⠉⠻⣿⣿⣿⣷⠀⠀⠀⠙⠧⣚⠀⠒⠀⠇⠘⠒⣏⢿⠀⠀⣰⣿⣿⣿⠏⡍⠀⠀⠀⠀⣽⣧⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⢸⡇⠀⠀⠀⠀⠀⠀⠀⣰⡿⠀⠀⠀⠀⠀⠀⠀⠀⡀⣰⠟⠁⠀⠀⠀⠀⠀⠀⠀⠈⢿⣿⣿⣿⣦⠀⠀⠐⣌⠓⠤⠤⠼⢄⠤⠔⣫⠎⢀⡜⢀⣼⣿⣿⡿⢁⠀⠀⠻⣿⣿⣿⣷⣄⠀⠀⠱⣄⡉⠓⠒⠓⠚⢉⡱⠋⢀⣴⣿⣿⡿⢋⠜⠀⠀⠀⠀⠀⣽⣹⠂⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⣠⣟⣧⠀⠀⠀⠀⠀⠀⢰⡿⠁⠀⡄⣀⣀⣤⡴⠶⠞⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠻⣿⣿⣿⣿⣤⣀⠈⠑⠲⠤⠤⠤⠴⠛⣁⡄⣊⣤⣿⣿⣿⠟⠁⠈⠀⠀⠀⠈⢿⣿⣿⣿⣷⣦⣄⣀⠉⢑⣒⣒⣚⣁⣤⣶⣿⣿⣿⠟⠁⠈⠀⠀⠀⠀⠀⢄⣻⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⣿⠙⠛⠶⣤⡀⠀⡀⣴⣿⠃⣠⣾⣷⣿⣉⠄⠡⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠻⢿⣿⣿⣿⣿⣶⣶⣤⣤⣤⣬⣶⣶⣿⣿⣿⣿⠟⠁⠀⠀⠀⠀⠀⠀⠈⠢⣙⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠛⠁⠀⠀⠀⠀⠀⠀⠀⠀⣄⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠻⢷⣦⣀⠀⠉⠓⠒⠾⢿⣾⣿⢸⣿⡷⡸⢌⠐⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠠⢙⠻⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠟⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠑⠤⠉⠛⠻⠿⠿⠿⠿⠿⠟⠛⠋⢁⣀⣤⣬⣾⡓⠀⠀⠀⠀⠀⠰⣸⣿⡗⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠈⠉⠉⠒⠒⠒⠺⢯⠿⠁⢸⢿⣷⡱⢊⠔⡡⢀⠀⠀⠀⠀⠀⠀⢰⣿⣿⣦⣴⣤⣄⣀⣀⠀⠀⠀⠉⠀⠐⠀⠉⠍⠉⠉⢡⠈⣀⡀⠀⠀⢀⣀⣀⣀⣀⣀⣀⣀⣀⠀⠀⣀⠀⠀⠀⠀⠊⠁⠀⣀⣠⣤⡴⠶⣿⣿⡿⡿⠋⠉⠻⣍⠀⠀⠀⠀⢐⣿⢿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡴⡤⡴⡒⡤⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⣿⣿⡱⣍⠲⡐⠄⠂⠀⠀⠀⠀⠈⡻⠟⠉⠉⠓⢿⣯⡉⠉⠙⣿⠛⠿⠷⠶⠶⣶⣤⣤⣤⣠⣀⣀⣉⣀⣉⣁⣀⣀⣀⣀⣀⣀⣠⣤⣤⣤⣤⠤⠤⠶⠶⠖⢻⠛⠋⠉⣼⠇⣤⣿⠟⠉⠀⠀⠀⠀⠀⠀⠀⠀⠀⢬⣿⡎⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⢾⡟⠁⠀⠀⠁⠘⣆⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢿⣿⣗⢮⡓⣌⠒⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠳⢄⡀⣿⠁⠀⠀⠀⠀⢰⠀⠀⠀⠀⠈⠉⢨⡉⠉⠉⠉⠉⠉⢩⡍⠉⠁⠀⠀⠀⢠⠀⠀⠀⠀⠀⣾⠁⠀⢠⡿⢤⡿⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣽⢿⠇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⣿⡇⠀⠀⠀⠀⠀⡯⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⣿⣿⢶⡹⣄⠣⠀⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠑⢿⣧⣆⣀⣀⡀⣸⣀⠀⠀⠀⠀⠀⢸⠀⠀⠀⠀⠀⠀⣼⡀⠀⠀⠀⠀⢀⣼⣀⣀⣀⡤⣼⡗⠒⠛⢉⣷⠎⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣼⡟⡏⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢰⣿⠷⠤⠤⠠⠤⢾⠃⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠸⣿⣧⢷⡸⣑⠂⠄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠢⣄⠀⠉⢹⠋⠋⠙⠚⠒⠒⢻⠓⠒⠒⠒⠋⠙⢻⠋⠉⠉⠉⠉⠉⡏⠀⠀⠀⠀⠀⡇⠀⢀⠞⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣒⡿⡟⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣸⣿⠂⠀⠀⠀⠀⢸⠆⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢯⣿⣮⢷⡑⢎⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠓⢤⣸⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⠀⠀⠀⠀⠀⠀⠇⠀⠀⠀⠀⠀⣧⡴⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡄⣿⡳⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⠀⠀⠀⣀⠔⠛⡆⠀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⡈⢿⣿⣧⢛⣌⠢⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢦⡀⠀⠀⠀⠈⠓⢤⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠸⠀⠀⠀⠀⠀⠀⡀⠀⠀⣀⡤⠞⠁⠀⠀⠀⣠⡤⠀⠀⠀⠀⠀⠀⠀⠀⢀⣰⣿⡿⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⡀⠈⠉⠀⠀⠨⣻⡀⠀⠀⠀⠀⠀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⠓⠘⡿⣿⣳⣌⢣⠐⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡗⠤⠀⠀⠀⠀⠀⠈⠉⠓⠒⠦⢤⣀⣀⣀⣀⣀⣴⣀⣀⣀⣠⠤⠴⠷⠒⠋⠁⠀⠀⠀⠀⣠⣾⠟⠀⠀⠀⠀⠀⠀⠀⠀⣀⣶⣿⡻⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢰⣿⣿⣿⣷⣦⣠⣤⢤⣄⣟⠲⡤⢀⣀⡀⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⢿⢿⣷⣣⠘⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠳⣦⡄⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⣶⠿⠁⠀⠀⠀⠀⠀⠀⡀⠄⡑⣾⣟⠝⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣸⣿⣟⢧⣿⣿⠋⠀⠀⠀⠀⠀⠀⠀⠁⠉⢩⢶⡀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⢯⣿⣷⣫⠔⡡⠀⠂⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠻⣾⣦⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣀⣤⠾⠋⠁⠀⠀⠀⠀⠀⠀⡀⠢⢐⣸⣼⡿⠋⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⡶⢹⣿⡍⢾⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢯
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⢿⣿⣿⣌⡱⢈⠀⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠛⠷⣦⣔⣤⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⢀⣠⠴⠞⠋⠁⠀⠀⠀⠀⠀⠀⢀⠠⠐⢤⣡⣾⣯⠎⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣰⡽⠰⣿⣿⣖⡜⣿⣿⣧⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢄⣾
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⢿⣿⣷⣬⣒⠠⠁⠄⠠⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠉⠙⠛⠶⠶⠖⠶⠲⠲⠲⠐⠞⠉⠉⠉⠀⠀⠀⠀⠀⠀⠀⠀⠠⠀⢂⠄⣊⣵⡾⣻⠟⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⡇⠀⡟⢩⢋⣿⡟⢭⡉⠏⠙⠛⠛⠻⠻⠿⢭⣍⡹⣯⡔
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⠿⣳⣿⣷⡩⢆⡡⠐⡠⠐⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠁⠀⡀⠠⢀⠡⢀⠈⡂⡁⢆⠱⣌⣮⣟⡯⠎⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⡇⢸⣧⡳⣎⣿⡜⣦⡑⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠚⢳
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⠺⢟⣿⣶⡥⢣⡔⡡⢄⡡⢀⠄⡀⢄⡀⠀⠄⠀⢀⠀⠀⠠⠀⠠⠀⠀⠠⠀⠠⠀⠄⠁⠌⡀⠡⢀⠐⠠⢂⡐⠢⡐⢤⡱⣮⢷⡻⠟⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠻⣳⡸⣿⣷⢭⢻⣿⣶⣣⠄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣾
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠺⣿⣿⣾⣱⣦⡜⣤⢊⡔⠢⣄⠡⡈⠔⡠⢈⠄⡡⢈⠠⠈⡄⡁⠤⢁⡐⠌⡰⠠⢄⠡⢂⡌⣑⢢⣬⣵⣭⣷⠿⠛⠉⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠳⢽⣿⣎⣦⣽⣿⣿⣶⣤⣀⣀⣀⣀⣀⣀⣀⣀⣠⠛
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠁⠛⠽⢟⡶⢷⣮⣷⣤⣓⡱⢎⡴⣁⠎⡔⢢⣁⢣⡐⣌⠒⣤⢢⡱⣄⣣⣼⣬⡷⢿⣾⠯⠛⠋⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠹⣿⣿⠏⠀⠀⠀⠀⠀⠀⠀⠈⠉⠹⠻⡋⠁⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠉⠁⠚⠛⠿⠿⣻⢷⣻⣾⣟⣷⣮⣷⣝⡾⣿⣶⠷⠟⠚⠛⠓⠊⠉⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢹⣿⣃⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣷⠇⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⠺⣻⣆⣀⡀⠀⠀⠀⠀⠀⠀⢠⣽⠀⠀⠀
# ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠉⠉⠙⠋⠉⠉⠓⠉⠁⠀⠀⠀
